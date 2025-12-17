import pandas as pd
from openpyxl import load_workbook
from stocks_reader import read_and_print_rows
from utils import get_db_path, get_exec_path
from settings import Settings

def load_excel_data(file_path, sheet_name=None):
    """
    Loads the Excel file and returns the DataFrame.

    :param file_path: Path to the Excel file
    :param sheet_name: Name of the sheet to read (None for active)
    :return: Pandas DataFrame or None if error
    """
    try:
        wb = load_workbook(file_path)
        print(f"Workbook sheets: {wb.sheetnames}")
        print(f"Active sheet: {wb.active.title}")
        if sheet_name is None:
            sheet = wb.active
            sheet_name_for_pandas = wb.sheetnames.index(sheet.title)
        else:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found.")
            sheet = wb[sheet_name]
            sheet_name_for_pandas = sheet_name
        print(f"Reading sheet: {sheet.title}")
        print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")

        print("-" * 50)
        df = pd.read_excel(file_path, sheet_name=sheet_name_for_pandas, header=None, dtype=str)
        print(f"DataFrame shape: {df.shape}")
        print(f"Total rows: {len(df)}")
        print("-" * 50)
        return df
    except FileNotFoundError:
        print(f"File {file_path} not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def get_row_indices():
    row_input = input("Enter the row numbers or ranges (e.g., 1-5, 7, 10-12): ")
    row_indices = []
    for part in row_input.split(','):
        part = part.strip()
        if '-' in part:
            try:
                start_str, end_str = part.split('-', 1)
                start = int(start_str.strip())
                end = int(end_str.strip())
                if start <= end:
                    row_indices.extend(range(start, end + 1))
                else:
                    print(f"Invalid range: {part}")
            except ValueError:
                print(f"Invalid range: {part}")
        else:
            try:
                idx = int(part)
                row_indices.append(idx)
            except ValueError:
                print(f"Invalid number: {part}")
    row_indices = sorted(set(row_indices))
    return row_indices

def scan_for_markers(df, settings=Settings()):
    sections = {}
    for idx, row in df.iterrows():
        # Assume marker is in first column starts with '#'
        first_cell = str(row[0]).strip() if pd.notna(row[0]) else ''
        if first_cell.startswith('#'):
            marker = first_cell.strip()
            sections[marker] = idx + 1  # Store 1-based index
    if sections:
        print("Found the following markers:")
        for marker, row_num in sections.items():
            print(f"Marker: '{marker}' at row {row_num}")
        if '#FDB' in sections and '#FDE' in sections:
            print("Funding deposit marker found. loading fund data...")
            row_indices = list(range(sections['#FDB'] + 1, sections['#FDE']))
            print("Loading fund data rows:", row_indices)
            read_and_print_rows(df, 'deposit', row_indices, True, settings)
        if '#FWB' in sections and '#FWE' in sections:
            print("Funding withdraw marker found. loading fund data...")
            row_indices = list(range(sections['#FWB'] + 1, sections['#FWE']))
            print("Loading fund data rows:", row_indices)
            read_and_print_rows(df, 'withdraw', row_indices, True, settings)
        if '#TBB' in sections and '#TBE' in sections:
            print("Trade buy marker found. loading trade data...")
            row_indices = list(range(sections['#TBB'] + 1, sections['#TBE']))
            print("Loading trade buy data rows:", row_indices)
            read_and_print_rows(df, 'buy', row_indices, True, settings)
        if '#TSB' in sections and '#TSE' in sections:
            print("Trade sell marker found. loading trade data...")
            row_indices = list(range(sections['#TSB'] + 1, sections['#TSE']))
            print("Loading trade sell data rows:", row_indices)
            read_and_print_rows(df, 'sell', row_indices, True, settings)
        
    else:
        print("No markers found in the first column.")            
            

def main():
    default_path = get_exec_path( 'stocks_transactions.xlsx' )
    settings = Settings().load( get_exec_path( 'settings.json' ) )
    if settings is None:
        settings = Settings()  # Use default settings if loading fails
    file_path = input(f"Enter the path to the Excel file (default: {default_path}): ").strip()
    file_path = file_path if file_path else default_path
    sheet_name = input("Enter sheet name (leave blank for active sheet): ").strip()
    sheet_name = sheet_name if sheet_name else None
    try:
        df = load_excel_data(file_path, sheet_name)
        if df is not None:
            operation = input("Enter operation: (r)ead and save rows, (s)can for markers: ").strip().lower()
            if operation == 's':
                scan_for_markers(df, settings)
            else:
                section = input("Enter the section to save the data (e.g., 'deposit', 'withdraw', 'buy', 'sell'): ").strip().lower()
                
                if section == 'deposit' or section == 'withdraw':
                    print("Assumed column mapping for funds:")
                    print("2: trade_date, 3: source, amount_SAR, 4: amount_USD, 5: rate_exchange")
                elif section == 'buy' or section == 'sell':
                    print("Assumed column mapping for trades:")
                    print("2: trade_date, 3: symbol, 4: filled_qty, 5: price, 6: fees, 7: vat, 9: cost_value, 0: is_position_open (for buy)")
                else:
                    print("Invalid section.")
                    return                
                row_indices = get_row_indices()
                if row_indices:
                    read_and_print_rows(df, section, row_indices, True, settings)
                else:
                    print("No valid rows entered.")
        else:
            print("Failed to load data.")
    except KeyboardInterrupt:
        print("\nExiting...")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
